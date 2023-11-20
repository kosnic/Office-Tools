import openpyxl
TarWorb=openpyxl.load_workbook('C:\\Users\\DELL\\Desktop\\工作岗位搞科研\\副副本本2023桐乡工单信息表.xlsx')
SrcWorb=openpyxl.load_workbook('C:\\Users\\DELL\\Desktop\\工作岗位搞科研\\副本工单查询装机历史单列表.xlsx')
Tarws=TarWorb["在途"]
Srcws=SrcWorb["工单列表"]
CurRow=Tarws.max_row+1
Columns={'A':'Q','B':'L','C':'F','D':'T','E':'E','F':'AQ','G':'BB','H':'AD','K':'B','M':'A','P':'G','Q':'D','R':'H','S':'AC'}
for SrcRow in range(2,Srcws.max_row):
    for TaC,SrC in Columns.items():
        Tarws[f'{TaC}{CurRow}'].value=Srcws[f'{SrC}{SrcRow}'].value
    CurRow+=1
TarWorb.save('C:\\Users\\DELL\\Desktop\\工作岗位搞科研\\副副本本2023桐乡工单信息表.xlsx')
print('Done')
